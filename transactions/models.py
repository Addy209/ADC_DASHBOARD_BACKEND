from django.db import models
import datetime
import openpyxl as xl
import inspect
from django.dispatch import receiver
from django.shortcuts import get_object_or_404
# Create your models here.

class DailyTransaction(models.Model):
    date=models.DateField(primary_key=True)
    mb_fintxns=models.PositiveBigIntegerField()
    mb_nonfintxns=models.PositiveBigIntegerField()
    mb_totaltxn=models.PositiveBigIntegerField()
    mb_td=models.PositiveBigIntegerField()
    mb_td_percent=models.FloatField()
    mb_bd=models.PositiveBigIntegerField()
    upi_fintxns=models.PositiveBigIntegerField()
    upi_nonfintxns=models.PositiveBigIntegerField()
    upi_totaltxn=models.PositiveBigIntegerField()
    upi_td=models.PositiveBigIntegerField()
    upi_td_percent=models.FloatField()
    upi_bd=models.PositiveBigIntegerField()
    imps_totaltxn=models.PositiveBigIntegerField()
    imps_td=models.PositiveBigIntegerField()
    imps_td_percent=models.FloatField()
    imps_bd=models.PositiveBigIntegerField()

    class Meta:
        ordering=("date",)

    def __str__(self) -> str:
        return f"Transaction Counts for {self.date}"
    
    @classmethod
    def fifteendaydata(cls):
        end=cls.objects.latest('date').date
        start=end-datetime.timedelta(30)
        return cls.objects.filter(date__range=(start,end))
    
    @classmethod
    def todaydata(cls):
        return cls.objects.latest('date')

    @classmethod
    def getDatabetweenspecifieddates(cls, fromdate, todate):
        return cls.objects.filter(date__range=(fromdate,todate))

    @staticmethod
    def doEntry(val):
        try:
            dt=DailyTransaction()
            dt.date=val[0].date()
            dt.mb_fintxns=val[1]
            dt.mb_nonfintxns=val[2]
            dt.mb_totaltxn=val[3]
            dt.mb_td=val[4]
            dt.mb_td_percent=round(val[5],2)
            dt.mb_bd=val[6]
            dt.upi_fintxns=val[7]
            dt.upi_nonfintxns=val[8]
            dt.upi_totaltxn=val[9]
            dt.upi_td=val[10]
            dt.upi_td_percent=round(val[11],2)
            dt.upi_bd=val[12]
            dt.imps_totaltxn=val[13]
            dt.imps_td=val[14]
            dt.imps_td_percent=round(val[15],2)
            dt.imps_bd=val[16]
            dt.save()
        except Exception as e:
            raise ValueError("Incorrect Value Encountered|")

    @classmethod
    def createEntry(cls, file):
        try:
            datafields=[1,4,5,6,7,8,9,11,12,13,14,15,16,18,19,20,21]
            wb=xl.load_workbook(file['originFileObj'], data_only=True)
            sheet=wb.active
            for i in range(2,sheet.max_column+1):
                temp=[]
                for j in datafields:
                    temp.append(sheet.cell(j,i).value)
                res=None
                try:
                    res=DailyTransaction.objects.get(date=temp[0].date())
                except:
                    pass
                if res:
                    raise Exception("Records for this date already found. Please make changes from Admin Dashboard|")
                
                DailyTransaction.doEntry(temp)

            return True
        except Exception as e:
            raise e

class IncrementalUser(models.Model):
    date=models.DateField(primary_key=True)
    inc_upiUsers=models.BigIntegerField()
    inc_mbUsers=models.BigIntegerField()

    def __str__(self) -> str:
        return f'{self.date}=> MB: {self.inc_mbUsers} and UPI: {self.inc_upiUsers}'

    def createInc(self, date ,incmb, incupi):
        self.date=date
        self.inc_mbUsers=incmb
        self.inc_upiUsers=incupi
        self.save()

    @classmethod
    def getfifteenddaydata(cls):
        end=cls.objects.latest("date").date
        start=end-datetime.timedelta(30)
        return cls.objects.filter(date__range=(start,end)).order_by("date")

class TotalUser(models.Model):
    date=models.DateField(primary_key=True)
    mb=models.BigIntegerField()
    upi=models.BigIntegerField()

    def __str__(self):
        return f'Date: {self.date}=> MB: {self.mb} and UPI: {self.upi}'

    def saveData(self, date, mb, upi):
        val=None
        try:
            val=get_object_or_404(TotalUser, date=date)
            raise Exception("Data for this Date already found. Please use Admin Panel if you wish to update entry|")
        except Exception as e:
            if val:
                raise e
            if mb<0:
                mb=0
            if upi<0:
                upi=0
            self.date=date
            self.mb=mb
            self.upi=upi
            self.save()


    


# @receiver(post_save, sender=XLSheet)
# def saveTransactionData(sender, instance, created, **kwargs):
#     print("in signal")
#     if created:
#         file=instance.xl
#         datafields=[1,4,5,6,7,8,9,11,12,13,14,15,16,18,19,20,21]
#         wb=xl.load_workbook(file, data_only=True)
#         sheet=wb.active
#         print(sheet.max_column, sheet.max_row)
#         for i in range(2,sheet.max_column+1):
#             temp=[]
#             for j in datafields:
#                 temp.append(sheet.cell(j,i).value)
            
#             res=None
#             try:
#                 res=DailyTransaction.objects.get(date=temp[0].date())
#             except:
#                 pass
#             if res:
#                 raise Exception("Duplicate Entry Found")
                
#             DailyTransaction.doEntry(temp)
#     instance.delete()

